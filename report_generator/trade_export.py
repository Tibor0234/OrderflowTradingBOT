from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from global_services.data.provider import DataProvider
from global_services.events.bus import EventBus
from global_services.events.utils import EventBusMsgType
from trading.market_entities.trade import Trade
from trading.market_entities.utils import Side


class TradeExcelExporter:
    """Exports closed trades to an Excel workbook grouped by session pair."""

    HEADERS = ["Symbol", "Side", "Avg Entry Price", "Avg Close Price", "Invested Value (USD)", "Leverage", "Realized PnL (USD)", "Open Time", "Close Time", "Duration (s)"]
    INVALID_SHEET_CHARS = str.maketrans({char: "-" for char in "\\/*?:[]"})

    def __init__(self, report_directory: Path):
        """Initialize the exporter and subscribe to trade lifecycle events."""
        self.report_directory = report_directory
        self.session_counter = None
        self.session_trades: list[Trade] = []

        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)

        EventBus().subscribe(EventBusMsgType.TRADE_CLOSE, self._on_trade_close)
        EventBus().subscribe(EventBusMsgType.SESSION_PAIR_END, self._on_session_pair_end)
        EventBus().subscribe(EventBusMsgType.PROCESS_END, self._save_workbook)

    def set_session_counter(self, session_counter):
        """Set the session counter used to identify session sheets."""
        self.session_counter = session_counter
        return self

    def _on_trade_close(self, trade: Trade):
        """Store a closed trade for the current session pair."""
        self.session_trades.append(trade)

    def _on_session_pair_end(self):
        """Export the current session pair's trades to a dedicated worksheet."""
        if not self.session_trades:
            return

        symbol = DataProvider().get_symbol()
        session_number = self.session_counter.session if self.session_counter else len(self.workbook.sheetnames) + 1
        sheet_name = f"Session {session_number} - {symbol}".translate(self.INVALID_SHEET_CHARS)[:31]
        sheet = self.workbook.create_sheet(title=sheet_name)
        metadata_keys = list(dict.fromkeys(
            key for trade in self.session_trades for key in trade.metadata
        ))
        sheet.append(self.HEADERS + [f"Meta: {key}" for key in metadata_keys])

        for trade in self.session_trades:
            sheet.append([
                symbol,
                "BUY" if trade.side == Side.BUY else "SELL",
                float(trade.execution_price),
                float(trade.avg_close_price),
                float(trade.invested_value),
                float(trade.leverage),
                float(trade.realized_pnl),
                self._format_time(trade.open_time),
                self._format_time(trade.close_time),
                (trade.close_time - trade.open_time) / 1000,
                *[self._format_metadata_value(trade.metadata.get(key)) for key in metadata_keys],
            ])

        self.session_trades = []
        self._save_workbook()

    def _save_workbook(self):
        """Save the workbook to the configured report directory."""
        if not self.workbook.sheetnames:
            return
        self.workbook.save(self.report_directory / "trades.xlsx")

    @staticmethod
    def _format_time(timestamp_ms):
        """Convert a millisecond timestamp to a formatted datetime string."""
        return datetime.fromtimestamp(timestamp_ms / 1000).strftime("%Y-%m-%d %H:%M:%S")

    @staticmethod
    def _format_metadata_value(value):
        """Convert a trade metadata value to an Excel-compatible representation."""
        if isinstance(value, Decimal):
            return round(float(value), 5)
        if value is None or isinstance(value, (str, int, bool)):
            return value
        if isinstance(value, float):
            return round(value, 5)
        return str(value)
